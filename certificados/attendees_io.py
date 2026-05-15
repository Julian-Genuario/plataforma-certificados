"""Parsing helpers for the attendees import flow.

Accepts CSV, XLSX or pasted text and yields cleaned (full_name, email) rows
plus per-row error reports. Header detection is tolerant of common Spanish
and English column names.
"""
import csv
import io
import re

from django.core.validators import validate_email
from django.core.exceptions import ValidationError


NAME_HEADERS = {
    "nombre", "nombres", "nombre completo", "nombre y apellido",
    "apellido y nombre", "name", "full name", "full_name", "fullname",
    "participante", "alumno", "asistente",
}
EMAIL_HEADERS = {
    "email", "e-mail", "e mail", "mail", "correo",
    "correo electronico", "correo electrónico",
}

MAX_NAME_LEN = 200
MAX_ROWS = 5000


class ParseError(Exception):
    pass


def _norm_header(h):
    return (h or "").strip().lower()


def _is_email_like(value):
    if not value:
        return False
    return "@" in value and "." in value.split("@")[-1]


def _detect_columns(header_row):
    """Given a header row, return (name_idx, email_idx) or (None, None) if
    no recognizable headers were found.
    """
    name_idx = None
    email_idx = None
    for i, cell in enumerate(header_row):
        n = _norm_header(cell)
        if n in NAME_HEADERS and name_idx is None:
            name_idx = i
        elif n in EMAIL_HEADERS and email_idx is None:
            email_idx = i
    if name_idx is None and email_idx is None:
        return None, None
    return name_idx, email_idx


def _split_text_line(line):
    """Split a textarea line by comma / semicolon / tab."""
    for sep in [",", ";", "\t"]:
        if sep in line:
            parts = [p.strip() for p in line.split(sep)]
            return parts
    return [line.strip()]


def _clean_row(name_idx, email_idx, row):
    """Pull name and email from row using detected indexes, or fall back to
    'two columns where one looks like an email'.
    """
    if name_idx is not None or email_idx is not None:
        name = (row[name_idx].strip() if name_idx is not None and name_idx < len(row) else "")
        email = (row[email_idx].strip() if email_idx is not None and email_idx < len(row) else "")
        return name, email

    cells = [str(c).strip() for c in row if str(c).strip()]
    if len(cells) < 2:
        return cells[0] if cells else "", ""

    if _is_email_like(cells[0]) and not _is_email_like(cells[1]):
        return cells[1], cells[0]
    return cells[0], cells[1]


def _validate(name, email):
    if not name:
        return "Falta el nombre"
    if len(name) > MAX_NAME_LEN:
        return f"Nombre supera {MAX_NAME_LEN} caracteres"
    if not email:
        return "Falta el email"
    try:
        validate_email(email)
    except ValidationError:
        return "Email inválido"
    return None


def _parse_rows(rows):
    """Generic row parser. rows: iterable of list[str]. First row is treated
    as header IFF it contains a recognizable name/email column.
    Returns (clean: list[(name, email)], errors: list[(line_no, reason, raw)]).
    """
    rows = list(rows)
    if not rows:
        return [], []

    name_idx, email_idx = _detect_columns(rows[0])
    has_header = name_idx is not None or email_idx is not None
    data_rows = rows[1:] if has_header else rows
    start_line = 2 if has_header else 1

    clean = []
    errors = []
    for offset, row in enumerate(data_rows):
        line_no = start_line + offset
        if not any(str(c).strip() for c in row):
            continue
        name, email = _clean_row(name_idx, email_idx, row)
        err = _validate(name, email)
        raw = " | ".join(str(c) for c in row)
        if err:
            errors.append((line_no, err, raw))
            continue
        clean.append((name, email))
        if len(clean) >= MAX_ROWS:
            errors.append((line_no + 1, f"Límite de {MAX_ROWS} filas alcanzado, resto ignorado", ""))
            break
    return clean, errors


def parse_csv(file_obj):
    raw = file_obj.read()
    if isinstance(raw, bytes):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ParseError("No se pudo decodificar el archivo (codificación desconocida).")
    else:
        text = raw

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    return _parse_rows(reader)


def parse_xlsx(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ParseError("openpyxl no está instalado. Agregalo a requirements.")

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"No se pudo abrir el archivo XLSX: {exc}")

    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in r])
    return _parse_rows(rows)


def parse_text(text):
    lines = (text or "").splitlines()
    rows = [_split_text_line(line) for line in lines if line.strip()]
    return _parse_rows(rows)


def parse_uploaded_file(file_obj):
    """Dispatch on filename extension."""
    name = (getattr(file_obj, "name", "") or "").lower()
    if name.endswith(".xlsx"):
        return parse_xlsx(file_obj)
    if name.endswith(".csv") or name.endswith(".txt"):
        return parse_csv(file_obj)
    raise ParseError("Formato no soportado. Subí un archivo .csv o .xlsx.")
